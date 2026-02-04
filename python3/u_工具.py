# -*- coding: utf-8 -*-
# @Time    : 2026-02-04
# @PreTime : 2026-01-30
# @Author  : hlmio
import os
import shutil
import time
import datetime
import random
import math
import threading
import re
import traceback
import pathlib
import functools
import csv
from functools import wraps
from decimal import *


# region 未分类




# region flask_api

def flask_get输入参数(request, 参数名, 默认值=None, 参数在json的路径=None):
    try:
        rst = request.values.get(参数名)
        if not rst:
            if not 参数在json的路径:
                rst = getDictValue(request.json, 参数名)
            else:
                rst = getDictValue(request.json, 参数在json的路径)

        if not rst:
            raise Exception(f"未找到{参数名}参数")
        return rst
    except:
        return 默认值


# endregion


__lock_print = threading.Lock()

def print_加锁(*args, **kwargs):
    with __lock_print:
        print(*args, **kwargs)



# import ctypes
# def change_locals(frame, 修改表={}):
#     frame.f_locals.update(修改表)
#     ctypes.pythonapi.PyFrame_LocalsToFast(
#         ctypes.py_object(frame),
#         ctypes.c_int(0)
#     )


# endregion 未分类


# region 生成器
def 每x行取第y行_生成器类(x, y):
    行数 = -1 - (y - 1)
    while True:
        行数 += 1
        if 行数 % x == 0:
            yield True
        else:
            yield False


def 每x行取任意行_生成器类(x, 行编号=[]):
    if isinstance(行编号, int):
        行编号 = [行编号]
    行数 = -1
    while True:
        行数 += 1
        余数 = 行数 % x
        if (余数 + 1) in 行编号:
            yield True
        else:
            yield False

# endregion 生成器


# region 装饰器
# -- 关于初始化区，扫描到几个@就执行几次

# region --线程
from concurrent.futures import ThreadPoolExecutor
__线程池_装饰专用 = ThreadPoolExecutor(12)


def 线程模式_改(is_VIP=False, VIP_name=None):  # 这里的参数，是给装饰器的参数
    # region 装饰器的初始化区1
    # endregion
    def wrap(func):
        # region 装饰器的初始化区3
        # endregion
        @wraps(func)  # 复制原始函数信息，并保留下来
        def inner(*args, **kwargs):  # args和kwargs，是原始函数的参数；args是元祖，kwargs是字典

            # region 执行原始函数前
            # endregion

            if is_VIP:
                rst = threading.Thread(target=func, name=VIP_name, args=args, kwargs=kwargs)
                rst.start()
            else:
                rst = __线程池_装饰专用.submit(func, *args, **kwargs)  # 执行原始函数

            # region 执行原始函数后
            # endregion

            return rst

        # region 装饰器的初始化区4
        # endregion
        return inner

    # region 装饰器的初始化区2
    # endregion
    return wrap

def 线程模式(func):
    # region 装饰器的初始化区3
    # endregion
    @wraps(func)  # 复制原始函数信息，并保留下来
    def inner(*args, **kwargs):  # args和kwargs，是原始函数的参数；args是元祖，kwargs是字典

        # region 执行原始函数前
        # endregion

        rst = __线程池_装饰专用.submit(func, *args, **kwargs)  # 执行原始函数

        # region 执行原始函数后
        # endregion

        return rst

    # region 装饰器的初始化区4
    # endregion
    return inner

# endregion 线程

# region --定时任务
try:
    from apscheduler.executors.pool import ThreadPoolExecutor, ProcessPoolExecutor
    from apscheduler.schedulers.background import BackgroundScheduler
except:
    print("pip install APScheduler")

_scheduler = None
_定时任务列表 = []


def 定时任务_注册(触发器类型='interval', id=None, 首次是否执行=True, *任务args, **任务kwargs):  # 这里的参数，是给装饰器的参数
    # region 装饰器的初始化区：（1）装饰了别人才会执行 （2）有几个@，就执行几次
    # endregion

    def wrap(func):
        @wraps(func)  # 复制原始函数信息，并保留下来
        def inner(*args, **kwargs):  # args和kwargs，是原始函数的参数；args是元祖，kwargs是字典

            # region 执行原始函数前
            注册定时任务(func, 触发器类型, args, kwargs, id, *任务args, **任务kwargs)
            # endregion

            if 首次是否执行:
                func(*args, **kwargs)

            # region 执行原始函数后
            # endregion

        return inner

    return wrap


def 定时任务_启动():  # 这里的参数，是给装饰器的参数
    # region 装饰器的初始化区：（1）装饰了别人才会执行 （2）有几个@，就执行几次
    # endregion

    def wrap(func):
        @wraps(func)  # 复制原始函数信息，并保留下来
        def inner(*args, **kwargs):  # args和kwargs，是原始函数的参数；args是元祖，kwargs是字典

            # region 执行原始函数前
            启动定时任务()
            # endregion

            func(*args, **kwargs)  # 执行原始函数

            # region 执行原始函数后
            # endregion

        return inner

    return wrap


def 注册定时任务(func, 触发器类型, args, kwargs, id, *任务args, **任务kwargs):
    新任务 = {}
    新任务["任务"] = func
    新任务["触发器类型"] = 触发器类型
    新任务["原始函数args"] = args
    新任务["原始函数kwargs"] = kwargs
    新任务["id"] = id
    新任务["任务args"] = 任务args
    新任务["任务kwargs"] = 任务kwargs
    global _定时任务列表
    _定时任务列表.append(新任务)


def 启动定时任务():
    __executors = {
        'default': ThreadPoolExecutor(20),  # 线程池
        'processpool': ProcessPoolExecutor(5)  # 进程池
    }
    __job_defaults = {
        'coalesce': True,
        # 当有任务中途中断，后面恢复后，有N个任务没有执行 coalesce：true ，恢复的任务会执行一次  coalesce：false，恢复后的任务会执行N次配合misfire_grace_time使用
        'max_instances': 1,  # 同一任务的运行实例个数
        'misfire_grace_time': 60  # 超时间隔，超过了就弃掉任务
    }
    global _scheduler
    _scheduler = BackgroundScheduler(executors=__executors, job_defaults=__job_defaults,
                                     timezone='Asia/Shanghai')
    for i in _定时任务列表:
        _scheduler.add_job(i["任务"], i["触发器类型"], i["原始函数args"], i["原始函数kwargs"], i["id"], *i["任务args"], **i["任务kwargs"])
    _scheduler.start()


@定时任务_启动()
def 启动定时任务_阻塞主线程():
    while True:
        time.sleep(60 * 60 * 1)


# endregion 定时任务

# endregion



# region 12.文件传输 upload
try:
    import paramiko
except:
    print("pip install paramiko")
# 通过ssh上传指定文件
def upload_by_ssh(locate文件全路径, remote文件全路径, host, port, username, password):
    try:
        t = paramiko.Transport((host, int(port)))
        t.connect(username=username, password=password)
        sftp = paramiko.SFTPClient.from_transport(t)
        try:
            sftp.put(locate文件全路径, remote文件全路径)
        except Exception as e:
            sftp.mkdir(os.path.dirname(remote文件全路径))
            sftp.put(locate文件全路径, remote文件全路径)
    finally:
        try:
            t.close()
        except: pass


from ftplib import FTP
def __ftp_upload(f, local_path, remote_path):
    fp = open(local_path, "rb")
    buf_size = 1024
    f.storbinary("STOR {}".format(remote_path), fp, buf_size)
    fp.close()

def __ftp_download(f, local_path, remote_path):
    fp = open(local_path, "wb")
    buf_size = 1024
    f.retrbinary('RETR {}'.format(remote_path), fp.write, buf_size)
    fp.close()


# 将本地文件上传到ftp,并且重命名为指定文件名
def upload_by_ftp(locate文件全路径, remote文件全路径, host, port, username, password):
    ftp = FTP()
    try:
        ftp.connect(host, int(port))
        ftp.login(username, password)
        __ftp_upload(ftp, locate文件全路径, remote文件全路径)
    finally:
        try:
            ftp.quit()
        except: pass

# endregion 12.文件传输


# region 11.excel
try:
    import xlrd
    from xlrd import xldate_as_datetime, xldate_as_tuple
except:
    print("pip install xlrd==1.2.0")
try:
    import openpyxl
    from openpyxl.utils import get_column_letter, column_index_from_string
except:
    print("pip install openpyxl")


excel类型 = {
    "xlrd": {
        "workbook": "<class 'xlrd.book.Book'>",
        "sheet": "<class 'xlrd.sheet.Sheet'>",
    },
    "openpyxl": {
        "workbook": "<class 'openpyxl.workbook.workbook.Workbook'>",
        "sheet": "<class 'openpyxl.worksheet.worksheet.Worksheet'>",
    }
}


def get_file_lines(文件全路径, txt_分隔符=",", excel_sheet下标或名称=0, encoding="utf8", txt_is去掉所有空行=True, is全部读取为字符串=True):
    rows = get_file_rows(文件全路径, txt_分隔符=txt_分隔符, excel_sheet下标或名称=excel_sheet下标或名称, encoding=encoding, txt_is去掉所有空行=txt_is去掉所有空行, is全部读取为字符串=is全部读取为字符串)

    col_names = rows[0]
    # 处理字段名缺失
    递增数 = 0
    for i in range(len(col_names)):
        if not col_names[i]:
            递增数 += 1
            col_names[i] = f'''f{递增数}'''

    # 处理字段名重复
    def add_suffix_to_duplicates(lst):
        count_dict = {}
        seen = set()
        result = []

        for item in lst:
            original_item = item
            if item in seen:
                count = count_dict.get(item, 0)
                count += 1
                new_item = f"{item}{count}"
                while new_item in seen:
                    count += 1
                    new_item = f"{item}{count}"
                count_dict[original_item] = count
            else:
                seen.add(item)
                new_item = item

            result.append(new_item)
            seen.add(new_item)

        return result
    col_names = add_suffix_to_duplicates(col_names)

    lines = []
    for row in rows[1:]:
        r_dict = {}
        for i, col in enumerate(row):
            r_dict[col_names[i]] = col
        lines.append(r_dict)
    return lines

def get_file_rows(文件全路径, txt_分隔符=",", excel_sheet下标或名称=0, encoding="utf8", txt_is去掉所有空行=True, is全部读取为字符串=True, newline=None, txt_is_csv=None):
    rows = []
    if "xls" in 文件全路径.lower() or "xlsx" in 文件全路径.lower():
        rows = _get_file_rows__excel(文件全路径, excel_sheet下标或名称, encoding)
    elif ("csv" in 文件全路径.lower() or txt_is_csv) and not (txt_is_csv is not None and not txt_is_csv):
        rows = _get_file_rows__csv(文件全路径, txt_分隔符, encoding, txt_is去掉所有空行, newline)
    else:
        rows = _get_file_rows__txt(文件全路径, txt_分隔符, encoding, txt_is去掉所有空行, newline)
    if is全部读取为字符串:
        new_rows = []
        for i in rows:
            new_rows.append(stream(i).map(lambda x:str(x)).collect())
        rows = new_rows
    return rows

def _get_file_rows__csv(文件全路径, 分隔符=",", encoding="utf8", is去掉所有空行=True, newline=None):
    rows = []
    with open(文件全路径, encoding=encoding, newline=newline) as csv_file:
        csv_reader = csv.reader(csv_file, delimiter=分隔符)
        for row in csv_reader:
            if is去掉所有空行 and not row:
                continue
            rows.append(row)
    # 处理bom
    bomList = [r"\ufeff", r"\ufffe"]
    for bom in bomList:
        if rows and rows[0] and rows[0][0].startswith(bom):
            rows[0][0] = rows[0][0].split(bom, 1)[1]
    return rows

def _get_file_rows__txt(文件全路径, 分隔符=",", encoding="utf8", is去掉所有空行=True, newline=None):
    if 分隔符 == ",":
        return _get_file_rows__csv(文件全路径, 分隔符, encoding, is去掉所有空行)
    rows = []
    with open(文件全路径, "r", encoding=encoding, newline=newline) as f:
        for line in f.readlines():
            line = line.rstrip("\n")
            if is去掉所有空行 and not line:
                continue
            row = line.split(分隔符)
            rows.append(row)
    # 处理bom
    bomList = [r"\ufeff", r"\ufffe"]
    for bom in bomList:
        if rows and rows[0] and rows[0][0].startswith(bom):
            rows[0][0] = rows[0][0].split(bom, 1)[1]
    return rows

def _get_file_rows__excel(文件全路径, sheet下标或名称=0, encoding="utf8"):
    rows = []
    sheet = get_excel_sheet(文件全路径, sheet下标或名称)
    for i in range(get_excel_行数(sheet)):
        row = get_excel_行(sheet, i)
        rows.append(row)
    return rows


def from_rows_to_excel(rows, 文件全路径="result.xlsx", sheetName=None):
    wb = openpyxl.Workbook(write_only=True)
    try:
        # sheet = wb.active
        sheet = wb.create_sheet()
        if sheetName:
            sheet.title = sheetName

        one_sheet_rows = rows
        groups = []
        单sheet数据上限 = 1000000
        if len(rows) > 单sheet数据上限:
            groups = split_list_by_count(rows, 单sheet数据上限)
            one_sheet_rows = groups[0]

        for i in one_sheet_rows:
            sheet.append(i)
        if len(groups) > 1:
            for group in groups[1:]:
                sheet = wb.create_sheet()
                for i in group:
                    sheet.append(i)
        rm(文件全路径)
        mkdir(get文件所在目录(文件全路径))
        wb.save(文件全路径)
    finally:
        wb.close()

def get_excel_workbook(文件路径, 底层实现="xlrd"):
    底层实现 = 底层实现.lower()
    if 底层实现.lower() == "xlrd":
        return _get_excel_workbook__xlrd(文件路径)
    if 底层实现.lower() == "openpyxl":
        return _get_excel_workbook__openpyxl(文件路径)
    return Exception("底层实现未支持")


def _get_excel_workbook__xlrd(文件路径):
    return xlrd.open_workbook(文件路径)


def _get_excel_workbook__openpyxl(文件路径):
    return openpyxl.load_workbook(文件路径)


def get_excel_sheet(文件路径, sheet下标或名称=0, 底层实现="xlrd"):
    底层实现 = 底层实现.lower()
    if 底层实现 == "xlrd":
        return _get_excel_sheet__xlrd(文件路径, sheet下标或名称)
    if 底层实现 == "openpyxl":
        return _get_excel_sheet__openpyxl(文件路径, sheet下标或名称)
    return Exception("底层实现未支持")


def _get_excel_sheet__xlrd(文件路径, sheet下标或名称=0):
    workbook = _get_excel_workbook__xlrd(文件路径)
    if isinstance(sheet下标或名称, str):
        return workbook.sheet_by_name(sheet下标或名称)
    else:
        return workbook.sheet_by_index(sheet下标或名称)


def _get_excel_sheet__openpyxl(文件路径, sheet下标或名称):
    workbook = _get_excel_workbook__openpyxl(文件路径)
    if isinstance(sheet下标或名称, str):
        return workbook.get_sheet_by_name(sheet下标或名称)
    else:
        return workbook.worksheets[sheet下标或名称]


def get_excel_行数(sheet):
    def xlrd_sheet():
        return sheet.nrows

    def openpyxl_sheet():
        return sheet.max_row

    def default():
        raise Exception("参数类型未支持")

    switch = {
        excel类型["xlrd"]["sheet"]: xlrd_sheet,
        excel类型["openpyxl"]["sheet"]: openpyxl_sheet,
    }
    return switch.get(repr(type(sheet)), default)()


def get_excel_列数(sheet):
    def xlrd_sheet():
        return sheet.ncols

    def openpyxl_sheet():
        return sheet.max_column

    def default():
        raise Exception("参数类型未支持")

    switch = {
        excel类型["xlrd"]["sheet"]: xlrd_sheet,
        excel类型["openpyxl"]["sheet"]: openpyxl_sheet,
    }
    return switch.get(repr(type(sheet)), default)()


def get_excel_值(sheet, 行下标, 列下标):
    def xlrd_sheet():
        return _get_excel_合并单元格__xlrd(sheet, 行下标, 列下标)

    # 待处理合并单元格
    def openpyxl_sheet():
        return sheet.cell(row=行下标 + 1, column=列下标 + 1).value

    def default():
        raise Exception("参数类型未支持")

    switch = {
        excel类型["xlrd"]["sheet"]: xlrd_sheet,
        excel类型["openpyxl"]["sheet"]: openpyxl_sheet,
    }
    return switch.get(repr(type(sheet)), default)()


def get_excel_值_by序号(sheet, 行序号, 列序号):
    行下标 = to_excel序号_数字(行序号) - 1
    列下标 = to_excel序号_数字(列序号) - 1
    return get_excel_值(sheet, 行下标, 列下标)


def get_excel_值_by单词(sheet, 单词="A1"):
    单词match = re.match("([a-z A-Z]*)([0-9]*)([a-z A-Z]*)", 单词)
    列序号 = 单词match.group(1)
    if not 列序号: 列序号 = 单词match.group(3)
    行下标 = int(单词match.group(2)) - 1
    列下标 = to_excel序号_数字(列序号) - 1
    return get_excel_值(sheet, 行下标, 列下标)


def get_excel_行(sheet, 行下标):
    def xlrd_sheet():
        值list = []
        for j in range(get_excel_列数(sheet)):
            值 = get_excel_值(sheet, 行下标, j)
            值list.append(值)
        return 值list

    def default():
        raise Exception("参数类型未支持")

    switch = {
        excel类型["xlrd"]["sheet"]: xlrd_sheet,
    }
    return switch.get(repr(type(sheet)), default)()


def get_excel_列(sheet, 列下标):
    def xlrd_sheet():
        值list = []
        for i in range(get_excel_行数(sheet)):
            值list.append(get_excel_值(sheet, i, 列下标))
        return 值list
        return sheet.ncols

    def default():
        raise Exception("参数类型未支持")

    switch = {
        excel类型["xlrd"]["sheet"]: xlrd_sheet,
    }
    return switch.get(repr(type(sheet)), default)()


def get_excel_表头(sheet, 表头行下标__int或list):
    def xlrd_sheet():
        表头行下标 = 表头行下标__int或list
        if isinstance(表头行下标, int):
            表头list = get_excel_行(sheet, 表头行下标)
            return stream(表头list).map(lambda i: str(i).strip()).collect()
        assert not isinstance(表头行下标, (int, str))
        sheet_columns = sheet.ncols
        表头list = ["" for i in range(sheet_columns)]
        表头行下标.sort()
        for i in 表头行下标:
            for j in range(sheet_columns):
                if _is_excel_合并单元格__xlrd(sheet, i, j):
                    if _is_excel_第一行的合并单元格__xlrd(sheet, i, j):
                        cell_value = get_excel_值(sheet, i, j)
                        表头list[j] = f"{表头list[j]}-{cell_value}"
                else:
                    cell_value = sheet.cell_value(i, j)
                    表头list[j] = f"{表头list[j]}-{cell_value}"
        表头list = stream(表头list).map(lambda i: i[1:]).collect()
        return 表头list

    def default():
        raise Exception("参数类型未支持")

    switch = {
        excel类型["xlrd"]["sheet"]: xlrd_sheet,
    }
    return switch.get(repr(type(sheet)), default)()


def to_excel序号_字母(数字):
    if isinstance(数字, str):
        try:
            数字 = int(数字)
        except Exception as e:
            return 数字
    return get_column_letter(数字)


def to_excel序号_数字(字母):
    if isinstance(字母, int): return 字母
    return column_index_from_string(字母)


def get_excel序号_列表(开头序号_字母或数字__包括开头, 结尾序号_字母或数字__包括结尾, 生成字母列表=True):
    开头序号 = to_excel序号_数字(开头序号_字母或数字__包括开头)
    结尾序号 = to_excel序号_数字(结尾序号_字母或数字__包括结尾)
    返回列表 = []
    for i in range(开头序号, 结尾序号 + 1):
        返回列表.append(i)
    if 生成字母列表:
        返回列表 = stream(返回列表).map(lambda i: to_excel序号_字母(i)).collect()
    return 返回列表


def _get_excel_合并单元格__xlrd(sheet, 行下标, 列下标):
    单元格值 = sheet.cell_value(行下标, 列下标)
    # 处理日期字段，格式化成 '2025-12-12 11:02:00'
    if sheet.cell(rowx=行下标, colx=列下标).ctype == xlrd.XL_CELL_DATE:
        单元格值 = xldate_as_datetime(单元格值, sheet.book.datemode).strftime("%Y-%m-%d %H:%M:%S")
    if not 单元格值:
        merged = sheet.merged_cells
        for (row_index_min, row_index_max, col_index_min, col_index_max) in merged:
            if row_index_min <= 行下标 and 行下标 < row_index_max:
                if col_index_min <= 列下标 and 列下标 < col_index_max:
                    单元格值 = sheet.cell_value(row_index_min, col_index_min)
                    break
    return 单元格值


def _is_excel_合并单元格__xlrd(sheet, 行下标, 列下标):
    merged = sheet.merged_cells
    for (row_index_min, row_index_max, col_index_min, col_index_max) in merged:
        if row_index_min <= 行下标 and 行下标 < row_index_max:
            if col_index_min <= 列下标 and 列下标 < col_index_max:
                return True
    return False


def _is_excel_第一行的合并单元格__xlrd(sheet, 行下标, 列下标):
    merged = sheet.merged_cells
    for (row_index_min, row_index_max, col_index_min, col_index_max) in merged:
        if row_index_min == 行下标:
            if col_index_min <= 列下标 and 列下标 < col_index_max:
                return True
    return False


# endregion 11.excel


# region 10.shell
import subprocess
import platform


def is_linux_system():
    return 'linux' in platform.system().lower()


def is_windows_system():
    return 'windows' in platform.system().lower()


def shell(cmd, stdout=subprocess.PIPE, encoding="utf8", shell=True, check=True, **kwargs):
    return subprocess.run(cmd, stdout=stdout, encoding=encoding, shell=shell, check=check, **kwargs) \
        .stdout


# endregion 10.shell


# region 9.配置相关
import configparser


def _configparser_to_dict(config):
    my_dict = dict(config._sections)
    for key in my_dict:
        my_dict[key] = dict(my_dict[key])
    return my_dict


class 配置类:
    @staticmethod
    def 实例化():
        return 配置类()

    def __init__(self):
        self.数据源 = {}
        self.配置 = {}
        self.变量集 = {}
        self.关联表 = {}

    def _数据源转dict(self):
        路径 = self.数据源["路径"]
        类型 = self.数据源["类型"]
        来源 = self.数据源["来源"]

        if (not 类型) or 类型 == "auto":
            类型 = os.path.splitext(路径)[-1][1:].lower()

        try:
            if 类型 == "json":
                if 来源 == "filesystem" or 来源 == "file_system" or 来源 == "file" or 来源 == "fs":
                    dict配置 = to_json_obj(pathlib.Path(路径).read_text(encoding='UTF-8'))
                else:
                    raise Exception("配置加载失败，来源不支持")
            else:
                config = configparser.ConfigParser()
                config.read(路径, encoding='UTF-8')
                dict配置 = _configparser_to_dict(config)
        except:
            dict配置 = {}

        return dict配置

    def _设置数据源(self, 路径, 类型, 来源):
        self.数据源["路径"] = 路径
        self.数据源["类型"] = 类型.lower()
        self.数据源["来源"] = 来源.lower()
        return self

    def _get_dict_keyLocation_list(self, mydict={}, 分隔符=".", 前缀="", rstList=[]):
        for i in mydict.keys():
            if not 前缀:
                临时前缀 = f"{i}"
            else:
                临时前缀 = f"{前缀}{分隔符}{i}"

            if isinstance(mydict[i], dict):
                self._get_dict_keyLocation_list(mydict[i], 分隔符, 临时前缀, rstList)
            else:
                rstList.append(临时前缀)

        return rstList

    def 加载(self, 路径="./配置文件.json", 类型="auto", 来源="filesystem"):
        self.配置 = self._设置数据源(路径, 类型, 来源)._数据源转dict()
        return self

    def 取值(self, key="", 分隔符="."):
        return getDictValue(self.配置, key, 分隔符)

    def 关联(self, 变量集, 关联表={}, 分隔符="."):
        if not 关联表:
            关联表 = {}
            keyLocation = self._get_dict_keyLocation_list(变量集, 分隔符=分隔符)
            for i in keyLocation:
                关联表[i] = i
        self.变量集 = 变量集
        self.关联表 = 关联表
        return self

    def 导出(self, is_del_before=False, vars={}):
        config_filePath = self.数据源["路径"]
        is_file_exist = os.path.exists(config_filePath)

        if is_file_exist:
            if is_del_before:
                os.remove(config_filePath)
            else:
                return self

        所在目录 = os.path.dirname(config_filePath)
        if 所在目录:
            if not os.path.exists(所在目录):
                os.makedirs(所在目录)

        if not vars:
            vars = self.变量集
        with open(config_filePath, 'w', encoding='utf-8') as f:
            json.dump(vars, f, ensure_ascii=False, indent=2)

        return self

    def 重载(self, 分隔符=".", is_override_vars=True):
        # 加载配置
        self.配置 = self._数据源转dict()

        if is_override_vars:
            # 覆写变量
            for key in self.关联表.keys():
                old_value = getDictValue(self.变量集, key, 分隔符=分隔符)
                now_value = getDictValue(self.配置, self.关联表.get(key), old_value, 分隔符)
                setDictValue(self.变量集, key, now_value)

        return self

    def load(self, 路径="./配置文件.json", 类型="auto", 来源="filesystem"):
        return self.加载(路径, 类型, 来源)

    def get(self, key="", 分隔符="."):
        return self.取值(key, 分隔符)

    def link(self, 变量集, 关联表={}, 分隔符="."):
        return self.关联(变量集, 关联表, 分隔符=分隔符)

    def reload(self, 分隔符=".", is_override_vars=True):
        return self.重载(分隔符, is_override_vars)

    def export(self, is_del_before=False, vars={}):
        return self.导出(is_del_before, vars)

    def all(self, 配置文件路径, 变量集, 类型="auto", 来源="filesystem", 关联表={}, 分隔符=".", is_del_before=False, export_vars={},
            is_override_vars=True):
        return self.load(配置文件路径, 类型, 来源).link(变量集, 关联表, 分隔符=分隔符).export(is_del_before, export_vars).reload(分隔符,
                                                                                                           is_override_vars)


配置 = 配置类.实例化()

# endregion 9.配置相关


# region 8.dao
try:
    import openpyxl
except:
    print("pip install openpyxl")


def __get_conf_vlaue(conf, key_list, default=""):
    value = default
    for key in key_list:
        try:
            value = conf[key]
            return value
        except:
            continue
    return value


# region --redis

try:
    import redis as redis_py
except:
    print("pip install redis")

try:
    from rediscluster import RedisCluster
except:
    print("pip install redis-py-cluster")

_redis_conf = {
    "host": "127.0.0.1",
    "port": 6375,
    "username": "",
    "password": "redis",
    "db": 0,

    "decode_responses": True,
    "charset": "utf-8",
    "startup_nodes": []
    # "startup_nodes": [
    #     {"host": "127.0.0.1", "port": 7001},
    #     {"host": "127.0.0.1", "port": 7002},
    # ]
}


def _get_redis_conf(new_conf={}):
    conf = {}
    conf["host"] = new_conf.get("host", _redis_conf["host"])
    conf["port"] = new_conf.get("port", _redis_conf["port"])
    conf["username"] = __get_conf_vlaue(new_conf, ["username", "user", "name", "userName"], _redis_conf["username"])
    conf["password"] = __get_conf_vlaue(new_conf, ["password", "pass", "pw"], _redis_conf["password"])
    conf["db"] = __get_conf_vlaue(new_conf, ["db", "database"], _redis_conf["db"])

    conf["decode_responses"] = __get_conf_vlaue(new_conf, ["decode_responses"], _redis_conf["decode_responses"])
    conf["charset"] = __get_conf_vlaue(new_conf, ["charset"], _redis_conf["charset"])
    conf["startup_nodes"] = new_conf.get("startup_nodes", _redis_conf["startup_nodes"])
    return conf


class Redis:
    def __init__(self, conf=_get_redis_conf()):
        if conf["startup_nodes"]:
            # redis集群
            self.conn = RedisCluster(startup_nodes=conf["startup_nodes"], password=conf["password"], decode_responses=conf["decode_responses"])
        else:
            # redis单点
            self.conn = redis_py.StrictRedis(host=conf["host"], port=conf["port"], password=conf["password"], db=conf["db"],
                                         decode_responses=conf["decode_responses"], charset=conf["charset"])

    def __del__(self):
        if self.conn:
            try:
                self.conn.close()
            except:
                pass

    @staticmethod
    def 实例化(new_conf={}):
        conf = _get_redis_conf(new_conf)
        return Redis(conf)

    def 分布式锁_加锁(self, 锁名, 加锁人, 超时时间_秒=30):
        rst = self.conn.set(name=锁名, value=加锁人, nx=True, ex=超时时间_秒)
        return rst

    def 分布式锁_解锁(self, 锁名, 加锁人):
        lua = f"""
            if redis.call('get', KEYS[1]) == ARGV[1] then 
                return redis.call('del', KEYS[1])
            else 
                return 0 
            end
        """
        cmd = self.conn.register_script(lua)
        rst = cmd(keys=[锁名], args=[加锁人])
        return rst


def init_redis(new_conf={}):
    return Redis.实例化(new_conf)

# endregion --redis

# region --mongo

try:
    import pymongo
except: pass


_mongo_conf = {
    "host": "127.0.0.1",
    "port": 27017,
    "username": "admin",
    "password": "admin",
    "db": "a1",
    "table": "a1",
}


def _get_mongo_conf(new_conf={}):
    conf = {}
    conf["host"] = new_conf.get("host", _mongo_conf["host"])
    conf["port"] = new_conf.get("port", _mongo_conf["port"])
    conf["username"] = __get_conf_vlaue(new_conf, ["username", "user", "name", "userName"], _mongo_conf["username"])
    conf["password"] = __get_conf_vlaue(new_conf, ["password", "pass", "pw"], _mongo_conf["password"])
    conf["db"] = __get_conf_vlaue(new_conf, ["db", "database"], _mongo_conf["db"])
    conf["table"] = __get_conf_vlaue(new_conf, ["table", "tab"], _mongo_conf["table"])
    return conf


class Mongo:
    def __init__(self, conf=_mongo_conf):
        self.conn = pymongo.MongoClient(f'''mongodb://{conf["username"]}:{conf["password"]}@{conf["host"]}:{conf["port"]}/''')

        self.is_ok = True
        self.count = 0
        self.lines = []
        # self.db
        # self.table
        self.use_db(conf["db"])
        self.use_table(conf["table"])

    def __del__(self):
        if self.conn:
            try:
                self.close()
            except:
                pass

    @staticmethod
    def 实例化(new_conf={}):
        conf = _get_mongo_conf(new_conf)
        return Mongo(conf)

    def close(self):
        self.conn.close()

    def use_db(self, name="admin"):
        self.db = self.conn[name]

    def use_table(self, name):
        self.table = self.db[name]

    def use(self, name, type="table"):
        if type == "db": return self.use_db(name)
        if type == "table": return self.use_table(name)

        return self.use_table(name)

    def get_table(self, name):
        return self.db[name]

    def __getitem__(self,name):
        return self.get_table(name)

    def exec(self, sql: str, params=None):
        if params:
            self.cursor.execute(sql, params)
        else:
            self.cursor.execute(sql)
        self.rows = self.cursor.fetchall()
        self.count = self.cursor.rowcount
        self.lines = self._rows_to_lines(self.rows, self.cursor)
        return self

    def call(self, proc_name: str, params=[]):
        self.cursor.callproc(proc_name, params)
        self.rows = self.cursor.fetchall()
        self.count = self.cursor.rowcount
        self.lines = self._rows_to_lines(self.rows, self.cursor)
        if params:
            select_params = ",".join([f'@_{proc_name}_{i}' for i in range(len(params))])
            self.cursor.execute(f"select {select_params}")
            in_out = self.cursor.fetchone()
            for i in range(len(params)):
                params[i] = in_out[i]
        return self

    def begin(self):
        self.conn.begin()
        return self

    def commit(self):
        self.conn.commit()
        return self

    def rollback(self):
        self.conn.rollback()
        return self


def init_mongo(new_conf={}):
    return Mongo.实例化(new_conf)

# endregion --mongo


# region --oracle
# https://blog.csdn.net/u013595395/article/details/108924071
try:
    import cx_Oracle
except:
    print("pip install cx-Oracle==8.3.0")

# os.environ['NLS_LANG'] = 'SIMPLIFIED CHINESE_CHINA.UTF8'
# select userenv('language') from dual;
os.environ['NLS_LANG'] = 'SIMPLIFIED CHINESE_CHINA.AL32UTF8'

_oracle_conf = {
    "host": "192.168.15.132",
    "port": 1521,
    "username": "c##dba",
    "password": "oracle",
    "db": "orcl"
}


def _get_oracle_conf(new_conf={}):
    conf = {}
    conf["host"] = new_conf.get("host", _oracle_conf["host"])
    conf["port"] = new_conf.get("port", _oracle_conf["port"])
    conf["username"] = __get_conf_vlaue(new_conf, ["username", "user", "name", "userName"], _oracle_conf["username"])
    conf["password"] = new_conf.get("password", _oracle_conf["password"])
    conf["db"] = new_conf.get("db", _oracle_conf["db"])
    return conf


class Oracle:
    def __init__(self, conf=_get_oracle_conf()):
        self.conn = cx_Oracle.connect(
            f'{conf["username"]}/{conf["password"]}@{conf["host"]}:{conf["port"]}/{conf["db"]}')
        self.cursor = self.conn.cursor()

        self.count = 0
        self.rows = []
        self.lines = []
        self.title = []

    def __del__(self):
        self.close()

    def close(self):
        if self.cursor:
            try:
                self.cursor.close()
            except: pass
        if self.conn:
            try:
                self.conn.close()
            except: pass

    @staticmethod
    def 实例化(new_conf={}):
        conf = _get_oracle_conf(new_conf)
        return Oracle(conf)


    def exec(self, sql: str, params=None):
        if params:
            rst = self.cursor.execute(sql, params)
        else:
            rst = self.cursor.execute(sql)

        if rst:
            cursor = self.cursor
            self.rows = cursor.fetchall()
            self.count = cursor.rowcount
            self.lines = self._rows_to_lines(self.rows, cursor)
            self.title = [c[0] for c in cursor.description]
        else:
            self.rows = []
            self.count = 0
            self.lines = []
            self.title = []
        return self

    def exec_many(self, insert_sql: str, insert_data_rows: list, batch_size=7000, 是否每批次都提交=True):
        rows = insert_data_rows
        for i in split_list_by_count(rows, batch_size):
            self.cursor.executemany(insert_sql, i)
            if 是否每批次都提交:
                self.commit()

    def call(self, proc_name: str, params=[]):
        in_out = self.cursor.callproc(proc_name, params)
        cur_index = -1;
        for i in range(len(params)):
            params[i] = in_out[i]
            if repr(type(in_out[i])) == "<class 'cx_Oracle.Cursor'>":
                cur_index = i

        if cur_index != -1 and in_out[cur_index]:
            cursor = in_out[cur_index]
            self.rows = cursor.fetchall()
            self.count = cursor.rowcount
            self.lines = self._rows_to_lines(self.rows, cursor)
            self.title = [c[0] for c in cursor.description]
        else:
            self.rows = []
            self.count = 0
            self.lines = []
            self.title = []
        return self

    def call_func(self, proc_name: str, params=[], 返回值类型=None):
        if not 返回值类型:
            返回值类型 = cx_Oracle.STRING
        rst = self.cursor.callfunc(proc_name, 返回值类型, params)

        cur_index = -1;
        for i in range(len(params)):
            if repr(type(params[i])) == "<class 'cx_Oracle.CURSOR'>":
                cur_index = i

        if cur_index != -1 and params[cur_index]:
            cursor = params[cur_index].getvalue()
            self.rows = cursor.fetchall()
            self.count = cursor.rowcount
            self.lines = self._rows_to_lines(self.rows, cursor)
            self.title = [c[0] for c in cursor.description]
        else:
            self.rows = []
            self.count = 0
            self.lines = []
            self.title = []
        return rst;


    def begin(self):
        self.conn.begin()
        return self

    def commit(self):
        self.conn.commit()
        return self

    def rollback(self):
        self.conn.rollback()
        return self


    def to_excel(self, sql:str, 文件全路径:str, 每次取多少行=7000):
        wb = openpyxl.Workbook(write_only=True)
        try:
            当前写入多少行 = 0
            单sheet上限 = 1000000

            cursor = self.cursor
            cursor.execute(sql)

            # sheet = wb.active
            sheet = wb.create_sheet()
            # 3. 写入表头（可选）
            col_names = tuple(c[0] for c in cursor.description) if cursor.description else ()
            sheet.append(col_names)  # 写入列名
            当前写入多少行 += 1


            # 4. 分批读取+写入
            def read_db_batch(cursor, batch_size):
                """分批读取数据库数据，生成器（节省内存）"""
                while True:
                    rows = cursor.fetchmany(batch_size)
                    if not rows:
                        break
                    yield rows

            for data_batch in read_db_batch(cursor, 每次取多少行):
                # 待写入多少行 = len(data_batch)
                待写入多少行 = 每次取多少行
                if 当前写入多少行 + 待写入多少行 > 单sheet上限:
                    # 新开一个sheet
                    sheet = wb.create_sheet()
                    当前写入多少行 = 0

                for row_data in data_batch:
                    sheet.append(row_data)  # append()是openpyxl批量写入的高效方式
                当前写入多少行 += 待写入多少行

            rm(文件全路径)
            mkdir(get文件所在目录(文件全路径))
            wb.save(文件全路径)
        finally:
            wb.close()

        return self

    def get_rows_with_title(self):
        rows_with_title = []
        rows_with_title.append(self.title)
        for i in self.rows:
            rows_with_title.append(i)
        return rows_with_title

    def _rows_to_lines(self, rows, cursor):
        lines = []
        try:
            col_names = [c[0] for c in cursor.description]
            for row in rows:
                r_dict = {}
                for i, col in enumerate(row):
                    r_dict[col_names[i]] = col
                lines.append(r_dict)
        except:
            pass
        return lines


def oracle(new_conf={}):
    return Oracle.实例化(new_conf)


# endregion --oracle

# region --mysql
try:
    import pymysql
except:
    print("pip install PyMySQL")

_mysql_conf = {
    "host": "106.13.231.168",
    "port": 3306,
    "user": "root",
    "password": "ans573KUR",
    "db": "mysql",
    "charset": "utf8"
}


def _get_mysql_conf(new_conf={}):
    # if not new_conf:
    #     return copy.deepcopy(_mysql_conf)
    conf = {}
    conf["host"] = new_conf.get("host", _mysql_conf["host"])
    conf["port"] = new_conf.get("port", _mysql_conf["port"])
    conf["user"] = new_conf.get("user", _mysql_conf["user"])
    conf["password"] = new_conf.get("password", _mysql_conf["password"])
    conf["db"] = new_conf.get("db", _mysql_conf["db"])
    conf["charset"] = new_conf.get("charset", _mysql_conf["charset"])
    return conf


class Mysql:
    def __init__(self, conf=_mysql_conf):
        self.conn = pymysql.connect(host=conf["host"], port=conf["port"], user=conf["user"], password=conf["password"],
                                    db=conf["db"], charset=conf["charset"])
        self.cursor = self.conn.cursor()

        self.rows = []
        self.count = 0
        self.lines = []

        self.conn.autocommit(True)

    def __del__(self):
        if self.cursor:
            try:
                self.cursor.close()
            except:
                pass
        if self.conn:
            try:
                self.conn.close()
            except:
                pass

    @staticmethod
    def 实例化(new_conf={}):
        conf = _get_mysql_conf(new_conf)
        return Mysql(conf)

    def exec(self, sql: str, params=None):
        if params:
            self.cursor.execute(sql, params)
        else:
            self.cursor.execute(sql)
        self.rows = self.cursor.fetchall()
        self.count = self.cursor.rowcount
        self.lines = self._rows_to_lines(self.rows, self.cursor)
        return self

    def call(self, proc_name: str, params=[]):
        self.cursor.callproc(proc_name, params)
        self.rows = self.cursor.fetchall()
        self.count = self.cursor.rowcount
        self.lines = self._rows_to_lines(self.rows, self.cursor)
        if params:
            select_params = ",".join([f'@_{proc_name}_{i}' for i in range(len(params))])
            self.cursor.execute(f"select {select_params}")
            in_out = self.cursor.fetchone()
            for i in range(len(params)):
                params[i] = in_out[i]
        return self

    def begin(self):
        self.conn.begin()
        return self

    def commit(self):
        self.conn.commit()
        return self

    def rollback(self):
        self.conn.rollback()
        return self

    def _rows_to_lines(self, rows, cursor):
        try:
            # col_names = stream(self.cursor.description).map(lambda c: c[0]).collect()
            col_names = [c[0] for c in cursor.description]
        except:
            pass
        lines = []
        for row in rows:
            r_dict = {}
            for i, col in enumerate(row):
                r_dict[col_names[i]] = col
            lines.append(r_dict)
        return lines


def mysql(new_conf={}):
    return Mysql.实例化(new_conf)


# endregion --mysql

# region --postgresql
try:
    import psycopg2
except:
    print("pip install psycopg2-binary")

_pg_conf = {
    "host": "106.13.231.168",
    "port": 5432,
    "user": "root",
    "password": "ans573KUR",
    "db": "postgres",
    "charset": "utf8"
}


def _get_pg_conf(new_conf={}):
    conf = {}
    conf["host"] = new_conf.get("host", _mysql_conf["host"])
    conf["port"] = new_conf.get("port", _mysql_conf["port"])
    conf["user"] = new_conf.get("user", _mysql_conf["user"])
    conf["password"] = new_conf.get("password", _mysql_conf["password"])
    conf["db"] = new_conf.get("db", _mysql_conf["db"])
    conf["charset"] = new_conf.get("charset", _mysql_conf["charset"])
    return conf


class Pgsql:
    def __init__(self, conf=_pg_conf):
        self.conn = psycopg2.connect(host=conf["host"], port=conf["port"], user=conf["user"], password=conf["password"],
                                     database=conf["db"])
        self.cursor = self.conn.cursor()

        self.rows = []
        self.count = 0
        self.lines = []

        # self.conn.autocommit(True)

    def __del__(self):
        self.close()

    def close(self):
        if self.cursor:
            try:
                self.cursor.close()
            except:
                pass
        if self.conn:
            try:
                self.conn.close()
            except:
                pass

    @staticmethod
    def 实例化(new_conf={}):
        conf = _get_pg_conf(new_conf)
        return Pgsql(conf)

    def exec(self, sql: str, params=None):
        if params:
            self.cursor.execute(sql, params)
        else:
            self.cursor.execute(sql)
        self.rows = self.cursor.fetchall()
        self.count = self.cursor.rowcount
        self.lines = self._rows_to_lines(self.rows, self.cursor)
        return self

    def call(self, proc_name: str, params=[]):
        self.cursor.callproc(proc_name, params)
        self.rows = self.cursor.fetchall()
        self.count = self.cursor.rowcount
        self.lines = self._rows_to_lines(self.rows, self.cursor)
        return self


    def begin(self):
        self.conn.begin()
        return self

    def commit(self):
        self.conn.commit()
        return self

    def rollback(self):
        self.conn.rollback()
        return self


    def to_excel(self, sql:str, 文件全路径:str, 每次取多少行=7000):
        wb = openpyxl.Workbook(write_only=True)
        try:
            当前写入多少行 = 0
            单sheet上限 = 1000000

            cursor = self.cursor
            cursor.execute(sql)

            # sheet = wb.active
            sheet = wb.create_sheet()
            # 3. 写入表头（可选）
            col_names = tuple(c[0] for c in cursor.description) if cursor.description else ()
            sheet.append(col_names)  # 写入列名
            当前写入多少行 += 1


            # 4. 分批读取+写入
            def read_db_batch(cursor, batch_size):
                """分批读取数据库数据，生成器（节省内存）"""
                while True:
                    rows = cursor.fetchmany(batch_size)
                    if not rows:
                        break
                    yield rows

            for data_batch in read_db_batch(cursor, 每次取多少行):
                # 待写入多少行 = len(data_batch)
                待写入多少行 = 每次取多少行
                if 当前写入多少行 + 待写入多少行 > 单sheet上限:
                    # 新开一个sheet
                    sheet = wb.create_sheet()
                    当前写入多少行 = 0

                for row_data in data_batch:
                    sheet.append(row_data)  # append()是openpyxl批量写入的高效方式
                当前写入多少行 += 待写入多少行

            rm(文件全路径)
            mkdir(get文件所在目录(文件全路径))
            wb.save(文件全路径)
        finally:
            wb.close()

        return self


    def get_rows_with_title(self):
        rows_with_title = []
        col_names = [c[0] for c in self.cursor.description]
        rows_with_title.append(col_names)
        for i in self.rows:
            rows_with_title.append(i)
        return rows_with_title

    def _rows_to_lines(self, rows, cursor):
        try:
            # col_names = stream(self.cursor.description).map(lambda c: c[0]).collect()
            col_names = [c[0] for c in cursor.description]
        except:
            pass
        lines = []
        for row in rows:
            r_dict = {}
            for i, col in enumerate(row):
                r_dict[col_names[i]] = col
            lines.append(r_dict)
        return lines


def pgsql(new_conf={}):
    return Pgsql.实例化(new_conf)


# endregion --mysql

# endregion 8.dao


# region 7.to_xxx
import json
import uuid
import hashlib



try:
    import imgkit
    from HTMLTable import HTMLTable
except Exception as ex:
    print("pip install imgkit==1.2.3")
    print("pip install html-table==1.0")
# imgkit==1.2.3
# html-table==1.0
#
# imgkit底层调用wkhtmltopdf
# 需配置系统环境变量
# 或按以下例子主动写入配置
# config = imgkit.config(wkhtmltoimage='D:/a_soft/wkhtmltopdf/bin/wkhtmltoimage')
# imgkit.from_string(html_string, output_file, config=config)
def from_rows_to_img(写入的文件全路径, 数据行_rows, 表头行_row=None, title="", is删除写入的文件=False):
    table = HTMLTable(caption=title)

    if 表头行_row:
        # 表头行
        header_row = tuple(表头行_row)
        table.append_header_rows((
            header_row,
        ))

    table.append_data_rows(数据行_rows)

    # 标题样式
    table.caption.set_style({
        'font-size': '15px',
    })
    # 表格样式，即<table>标签样式
    table.set_style({
        'border-collapse': 'collapse',
        'word-break': 'keep-all',
        'white-space': 'nowrap',
        'font-size': '14px',
    })
    # 统一设置所有单元格样式，<td>或<th>
    table.set_cell_style({
        'width': "250px",
        'border-color': '#000',
        'border-width': '1px',
        'border-style': 'solid',
        'padding': '5px',
    })
    # 表头样式
    table.set_header_row_style({
        'color': '#fff',
        'background-color': '#48a6fb',
        'font-size': '18px',
    })

    # 覆盖表头单元格字体样式
    table.set_header_cell_style({
        'padding': '15px',
    })
    # 调小次表头字体大小
    # table[1].set_cell_style({
    #     'padding': '8px',
    #     'font-size': '15px',
    # })
    # 遍历数据行，如果增长量为负，标红背景颜色
    # for row in table.iter_data_rows():
    #     if row[2].value < 0:
    #         row.set_style({
    #             'background-color': '#ffdddd',
    #         })
    body = table.to_html()
    # html的charset='UTF-8'必须加上，否则中午会乱码
    html = "<!DOCTYPE html><html><head><meta charset='UTF-8'></head><body>{0}</body></html>".format(body)

    # 生成图片
    options = {
        "enable-local-file-access": None, # html携带图片有访问权限设置
        "width": 1355,
        "zoom": 2.1  # 越高像素越高
    }
    # 生成图片
    rm(写入的文件全路径)
    mkdir(get文件所在目录(写入的文件全路径))
    imgkit.from_string(html, 写入的文件全路径, options=options)
    with open(写入的文件全路径, 'rb') as f:
        img = f.read()
    if is删除写入的文件:
        rm(写入的文件全路径)
    return img



def from_hex_to_byte(str):
    return bytes.fromhex(str)
def from_byte_to_hex(字节):
    return 字节.hex()


# region 基本数据类型
def to_number(文本, isDecimal=True):
    if isDecimal:
        rst = Decimal("0.00")
    else:
        rst = 0.00

    try:
        if 文本:
            if not isinstance(文本, str):
                str(文本)
            if isDecimal:
                rst = Decimal(文本)
            else:
                rst = float(文本)
    except: pass
    return rst

def to_round(数据, 保留几位小数=0):
    数据_str = str(数据)
    整数位数 = 数据_str.find(".")
    if 数据 < 0:
        整数位数 -= 1
    return Context(prec=整数位数 + 保留几位小数, rounding=ROUND_HALF_UP).create_decimal(数据_str)


# endregion 基本数据类型



# region time  -- datetime.datetime是原点，是核心中间类

时间字符串_模板 = "%Y-%m-%d %H:%M:%S"

def from_oracle_time_format_to_py_time_format(format_str):
    # 定义映射关系（按长度从长到短，避免短匹配覆盖长匹配）
    mapping = {
        "YYYY": "%Y", "yyyy": "%Y",
        "MM": "%m", "mm": "%m",
        "DD": "%d", "dd": "%d",
        "HH24": "%H", "hh24": "%H",
        "MI": "%M", "mi": "%M",
        "SS": "%S", "ss": "%S",
        "FF6": "%f", "ff6": "%f",
        "FF3": "%f", "ff3": "%f",
        "FF": "%f", "ff": "%f"
    }

    # 按最长匹配优先排序
    pattern = re.compile("|".join(sorted(mapping.keys(), key=len, reverse=True)))

    # 替换函数
    def replace(match):
        return mapping[match.group(0)]

    return pattern.sub(replace, format_str)


def to_time_datetime(字符串or时间戳or时间元组=0, 格式字符串=时间字符串_模板, 增加几秒=0, 增加几分钟=0, 增加几小时=0, 增加几天=0, 是否由oracle格式转为py格式=True):
    obj = 字符串or时间戳or时间元组

    def from_str_to_datetime():
        字符串 = obj  # type:str
        字符串 = 字符串.strip()
        if 字符串 == "" or 字符串 == "0":
            return get_now_datetime()
        fmt = 格式字符串
        if not fmt:
            if len(字符串) == 10:
                fmt = "%Y-%m-%d"
            elif len(字符串) == 19 and "T" in 字符串:
                fmt = "%Y-%m-%dT%H:%M:%S"
            elif len(字符串) == 19:
                fmt = "%Y-%m-%d %H:%M:%S"
        if 是否由oracle格式转为py格式:
            fmt = from_oracle_time_format_to_py_time_format(fmt)
        return datetime.datetime.strptime(字符串, fmt)

    def from_时间元组_to_datetime():
        return datetime.datetime.fromtimestamp(obj)

    def from_普通元组_to_datetime():
        nonlocal obj
        普通元组 = obj  # type:tuple
        if 普通元组.count() < 9:
            补充个数 = 9 - 普通元组.count()
            for i in range(补充个数):
                普通元组 += (-1)
        obj = time.mktime(普通元组)
        return from_时间戳_to_datetime()

    def from_时间戳_to_datetime():
        时间戳 = obj  # type:float
        if 时间戳 == 0:
            return get_now_datetime()
        return datetime.datetime.fromtimestamp(时间戳)

    def from_datetime_to_datetime():
        return obj

    def default():
        raise Exception(f"参数类型未支持：{repr(type(obj))}")

    def get_now_datetime():
        return datetime.datetime.now()

    switch = {
        str: from_str_to_datetime,
        int: from_时间戳_to_datetime,
        float: from_时间戳_to_datetime,
        tuple: from_普通元组_to_datetime,
        time.struct_time: from_时间元组_to_datetime,
        datetime.datetime: from_datetime_to_datetime,
    }
    原点时间 = switch.get(type(obj), default)()

    # 接下来处理时间的增减
    增加的时间 = datetime.timedelta(seconds=增加几秒, minutes=增加几分钟, hours=增加几小时, days=增加几天)
    return 原点时间 + 增加的时间


def to_time_str(datetime_or_字符串or时间戳or时间元组=0, 格式字符串=时间字符串_模板, 增加几秒=0, 增加几分钟=0, 增加几小时=0, 增加几天=0, 格式字符串_旧="", 是否由oracle格式转为py格式=True):
    时间对象 = to_time_datetime(datetime_or_字符串or时间戳or时间元组, 格式字符串_旧, 增加几秒, 增加几分钟, 增加几小时, 增加几天, 是否由oracle格式转为py格式=是否由oracle格式转为py格式)
    if 是否由oracle格式转为py格式:
        格式字符串 = from_oracle_time_format_to_py_time_format(格式字符串)
    return 时间对象.strftime(格式字符串)


def to_time_unix(datetime_or_字符串or时间戳or时间元组=0, 增加几秒=0, 增加几分钟=0, 增加几小时=0, 增加几天=0):
    时间对象 = to_time_datetime(datetime_or_字符串or时间戳or时间元组, 时间字符串_模板, 增加几秒, 增加几分钟, 增加几小时, 增加几天)
    return time.mktime(时间对象.timetuple())


def to_time_tuple(datetime_or_字符串or时间戳or时间元组=0, 增加几秒=0, 增加几分钟=0, 增加几小时=0, 增加几天=0):
    时间对象 = to_time_datetime(datetime_or_字符串or时间戳or时间元组, 时间字符串_模板, 增加几秒, 增加几分钟, 增加几小时, 增加几天)
    return 时间对象.timetuple()


def to_now_datetime():
    return to_time_datetime(0)


def to_now_str(格式字符串=时间字符串_模板, 是否由oracle格式转为py格式=True):
    return to_time_str(0, 格式字符串, 是否由oracle格式转为py格式=是否由oracle格式转为py格式)


def to_now_unix():
    return to_time_unix(0)


def to_now_tuple():
    return to_time_tuple(0)


to_datetime = functools.partial(to_time_datetime)
to_时间字符串 = functools.partial(to_time_str)
to_时间戳 = functools.partial(to_time_unix)
to_时间元组 = functools.partial(to_time_tuple)
to_unix = functools.partial(to_time_unix)
to_now_时间戳 = functools.partial(to_now_unix)
to_now_时间元组 = functools.partial(to_now_tuple)
to_now_时间字符串 = functools.partial(to_now_str)


def x分钟前的unix(分钟数=0):
    return to_time_unix(0, 增加几分钟=-分钟数)


def 整分钟数的当前时间(整多少分钟=30):
    return 整分钟数的指定时间(整多少分钟=整多少分钟)


def 整分钟数的指定时间(指定的时间=None, 整多少分钟=30):
    分钟间隔 = 整多少分钟
    if not 指定的时间:
        指定的时间 = to_now_datetime()
    else:
        指定的时间 = to_time_datetime(指定的时间)
    当前整点时间 = 指定的时间.replace(minute=0, second=0, microsecond=0)
    当前整点时间_加一小时 = to_time_datetime(当前整点时间, 增加几小时=1)
    拿来比较的时间 = 当前整点时间_加一小时
    while 拿来比较的时间 >= 当前整点时间:
        拿来比较的时间 = to_time_datetime(拿来比较的时间, 增加几分钟=-分钟间隔)
        if 指定的时间 >= 拿来比较的时间:
            return 拿来比较的时间


# 获取当前时间的字符串
def getCurrentDatetime_str(format_str="%Y-%m-%d %H:%M:%S"):
    return datetime.datetime.now().strftime(format_str)


def from_excel小数日期_to_时间字符串(日期天数小数or字符串, 格式字符串=时间字符串_模板):
    日期天数小数 = float(日期天数小数or字符串)
    excel_date = datetime.datetime(1899, 12, 30) + datetime.timedelta(days=日期天数小数)  # Excel日期从1900年1月1日开始，减去1是因为Excel将1900年1月1日视为第1天
    # return to_time_str(excel_date, 格式字符串=格式字符串)
    return excel_date.strftime(格式字符串)  # 将日期转换为字符串格式

# endregion time


def to_self(obj):
    return to_json_obj(to_json_str(obj))


# region --json
class _MyEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, datetime.datetime):
            return to_time_str(obj)
        if isinstance(obj, bytes):
            return obj.decode("utf8")
        raise Exception(f"{obj}  {repr(type(obj))}  不能被处理")


def _get_dict(obj):
    try:
        rst = dict(obj)
    except:
        rst = obj.__dict__
    return rst


def from_class_to_dict(obj):
    obj = _get_dict(obj)
    for i in obj:
        try:
            obj[i] = _get_dict(obj[i])
        except:
            pass
    return obj


def to_json_str(obj, check_class=True, separators=(',',':') , ensure_ascii=False):
    if check_class:
        try:
            obj_dict = obj.__dict__
            obj = from_class_to_dict(obj)
        except:
            pass
    return json.dumps(obj, ensure_ascii=ensure_ascii, cls=_MyEncoder, separators=separators)


def to_json_file(obj, 文件对象, ensure_ascii=False, indent=2):
    return json.dump(obj, 文件对象, ensure_ascii=ensure_ascii, indent=indent)


def to_json_obj(字符串or文件对象):
    def from_str_to_dict():
        return json.loads(字符串or文件对象)

    def from_file_to_dict():
        return json.load(字符串or文件对象)

    def default():
        raise Exception("参数类型未支持")

    switch = {
        "<class 'str'>": from_str_to_dict,
        "<class '_io.TextIOWrapper'>": from_file_to_dict,
    }
    return switch.get(repr(type(字符串or文件对象)), default)()


# endregion --json


def to_md5(data):
    type_str = repr(type(data))
    if type_str != "<class 'bytes'>" and type_str != "<class 'str'>":
        data = json.dumps(data)
    if repr(type(data)) == "<class 'str'>":
        data = data.encode('utf-8')
    md5 = hashlib.md5()
    md5.update(data)
    return md5.hexdigest()


def to_uuid(去除中横线=True, 使用随机数=True):
    if 使用随机数:
        id = uuid.uuid4()
    else:
        id = uuid.uuid1()
    id = str(id)
    if 去除中横线:
        id = id.replace("-", "")
    return id


__to_变量名__pattren = re.compile(r'[\W+\w+]*?to_变量名\((\w+)\)')
__to_变量名__变量名集 = []

def to_变量名(变量):
    global __to_变量名__变量名集
    if not __to_变量名__变量名集:
        __to_变量名__变量名集 = __to_变量名__pattren.findall(traceback.extract_stack(limit=2)[0][3])
    return __to_变量名__变量名集.pop(0)

# endregion 7.to_xxx


# region 6.fileSystem
import glob

try:
    import pyperclip
except: pass
def ctrl_c(text):
    pyperclip.copy(text)
def ctrl_v():
    return pyperclip.paste()


def exist(文件全路径):
    return os.path.exists(文件全路径)


def isdir(文件全路径):
    if exist(文件全路径):
        return os.path.isdir(文件全路径)
    else:
        if get文件后缀(文件全路径):
            return False
        else:
            return True


def pwd(文件全路径):
    return os.path.abspath(文件全路径)


def ls(文件全路径, 包含前缀=True, 选项=""):
    选项 = 选项.lower()
    is_deep = False
    if ("p" in 选项) or ("r" in 选项):
        is_deep = True

    # glob
    if ("*" in 文件全路径) or ("?" in 文件全路径):
        if "**" in 文件全路径:
            is_deep = True
        return ls_glob(文件全路径, is_deep=is_deep, 包含前缀=包含前缀)

    if not exist(文件全路径):
        return []
    if not isdir(文件全路径):
        return [文件全路径]

    if is_deep:
        filePaths = _getAllFilePaths(文件全路径, is_deep=True)
        if not 包含前缀:
            filePaths = stream(filePaths).map(lambda i: get文件名(i)).collect()
        return filePaths
    else:
        if 包含前缀:
            return stream(os.listdir(文件全路径)) \
                .map(lambda i: os.path.join(文件全路径, i)).collect()
        else:
            return os.listdir(文件全路径)


def ls_glob(路径名称_带匹配符, is_deep=True, 包含前缀=True):
    filePaths = glob.glob(路径名称_带匹配符, recursive=is_deep)
    if not 包含前缀:
        filePaths = stream(filePaths).map(lambda i: get文件名(i)).collect()
    return filePaths


def mkdir(文件全路径, 选项="-p"):
    选项 = 选项.lower()
    if 文件全路径 and not exist(文件全路径):
        if ("p" in 选项) or ("r" in 选项):
            os.makedirs(文件全路径)
        else:
            os.mkdir(文件全路径)


def mk(文件全路径, 已有跳过_不删除=True, 选项="-p"):
    选项 = 选项.lower()
    if exist(文件全路径):
        if 已有跳过_不删除:
            return
        else:
            rm(文件全路径, "-rf")

    if isdir(文件全路径):
        mkdir(文件全路径, 选项)
    else:
        所在目录 = get文件所在目录(文件全路径)
        if 所在目录 and (not exist(所在目录)):
            mkdir(所在目录, 选项)
        with open(文件全路径, "a"):
            pass


def rm(文件全路径, 选项="-rf"):
    if exist(文件全路径):
        if isdir(文件全路径):
            if ("p" in 选项) or ("r" in 选项):
                shutil.rmtree(文件全路径)
            else:
                try:
                    os.rmdir(文件全路径)
                except:
                    stream(ls(文件全路径)).filter(lambda i: not isdir(i)) \
                        .forEach(lambda i: rm(i))
        else:
            os.remove(文件全路径)


def clear(文件全路径, 选项="-rf"):

    if not isdir(文件全路径): rm(文件全路径, 选项); return;
    if not exist(文件全路径): mk(文件全路径, 选项="-p"); return;
    stream(ls(文件全路径)).forEach(lambda f: rm(f, 选项))


def cp(旧文件, 新文件, 不删旧文件=True):
    旧文件类型 = "dir" if isdir(旧文件) else "file"
    新文件类型 = "dir" if isdir(新文件) else "file"

    # 确保文件夹存在
    if 新文件类型 == "dir":
        mk(新文件)
    if not exist(get文件所在目录(新文件)):
        mk(get文件所在目录(新文件))

    def file_file():
        # shutil.copyfile(旧文件,新文件)  # 只复制内容
        # 复制内容和权限 新文件不存在：新建，存在：覆盖
        shutil.copy(旧文件, 新文件)

    def file_dir():
        if not exist(新文件):
            mk(新文件)
        shutil.copy(旧文件, 新文件)

    def dir_file():
        if not exist(新文件):
            mk(新文件)
        with open(新文件, "ab") as ff:
            for i in ls(旧文件, 包含前缀=True):
                with open(i, "rb") as f:
                    ff.write(f.read())

    def dir_dir():
        shutil.copytree(旧文件, 新文件)

    def default():
        raise Exception("复制失败,参数类型未支持")

    switch = {
        "file-file": file_file,
        "file-dir": file_dir,
        "dir-file": dir_file,
        "dir-dir": dir_dir
    }
    switch.get(f"{旧文件类型}-{新文件类型}", default)()

    if not 不删旧文件:
        rm(旧文件, "-rf")


def get文件名(文件全路径):
    return os.path.basename(文件全路径)


def get文件后缀(文件全路径):
    return os.path.splitext(文件全路径)[1]


def get文件所在目录(文件全路径):
    return os.path.dirname(文件全路径)




def _getAllFilePaths(baseFilePath, is_deep=True):
    return getDeepFilePaths(baseFilePath, "*", is_deep)


# 递归获取 指定目录下，拥有指定后缀，的文件路径
def getDeepFilePaths(baseFilePath, ext_list="txt", is_deep=True):
    rst_filePaths = []
    _getDeepFilePaths(rst_filePaths, baseFilePath, ext_list, is_deep)
    return rst_filePaths


def _getDeepFilePaths(rst_filePaths, baseFilePath, ext_list="txt", is_deep=True):
    rst_filePaths += _getCurrentFilePaths(baseFilePath, ext_list)
    # 递归当前目录下的目录
    if is_deep:
        f_list = stream(os.listdir(baseFilePath)) \
            .map(lambda fileName: os.path.join(baseFilePath, fileName)) \
            .collect()
        stream(f_list) \
            .filter(lambda f: os.path.isdir(f)) \
            .forEach(lambda dir: _getDeepFilePaths(rst_filePaths, dir, ext_list, True))


def _getCurrentFilePaths(baseFilePath, ext_list="txt"):
    rst_filePaths = []
    if not baseFilePath:
        baseFilePath = "."
    # 处理ext后缀
    is_all_ext = False
    if not isinstance(ext_list, (list, tuple)):
        ext_list = [ext_list]
    selectExt_list = stream(ext_list).map(lambda i: i if (i and i[0] == ".") else f".{i}").collect()
    if ("." in selectExt_list) or (".None" in selectExt_list):
        selectExt_list.append("")
    if (".*" in selectExt_list):
        is_all_ext = True
    selectExt_list = stream(selectExt_list).filter(lambda i: i != "." and i != ".None" and i != ".*").collect()

    # 获取当前目录下的所有文件名
    f_list = stream(os.listdir(baseFilePath)) \
        .map(lambda fileName: os.path.join(baseFilePath, fileName)) \
        .collect()

    if is_all_ext:
        rst_filePaths += stream(f_list) \
            .filter(lambda f: not os.path.isdir(f)) \
            .collect()
    else:
        # 将当前目录下后缀名为指定后缀的文件，放入rst_filePaths列表
        stream(f_list) \
            .filter(lambda f: not os.path.isdir(f)) \
            .filter(lambda f: os.path.splitext(f)[1] in selectExt_list) \
            .forEach(lambda f: rst_filePaths.append(f))
    return rst_filePaths

# endregion 6.fileSystem


# region 5.线程序号

class 线程序号类:
    def __init__(self, 线程间独立=True):
        self.线程间独立 = 线程间独立
        self._序号集 = {}

    @staticmethod
    def 实例化(线程间独立=True):
        return 线程序号类(线程间独立)

    def 序号_重置(self, 下一个序号数字=1):
        序号值 = 下一个序号数字 - 1
        if self.线程间独立:
            线程标识符 = threading.currentThread().ident
        else:
            线程标识符 = "全局"
        self._序号集[线程标识符] = 序号值

    def 序号(self, 字符串模板="(1)"):
        # region 获取序号值
        if self.线程间独立:
            线程标识符 = threading.currentThread().ident
        else:
            线程标识符 = "全局"

        try:
            序号值 = self._序号集[线程标识符]
        except:
            序号值 = 0

        try:
            序号值 += 1
        except:
            序号值 = 1
        self._序号集[线程标识符] = 序号值
        # endregion

        # region 输出序号值
        if repr(type(字符串模板)) == "<class 'int'>": return 序号值
        字符串模板 = str(字符串模板)
        数字match = re.match(r"([^0-9]*)([0-9]*)([\s\S]*)", 字符串模板)
        模板序号值 = f"{数字match.group(1)}{序号值}{数字match.group(3)}"
        return 模板序号值
        # endregion


_静态序号生成器 = 线程序号类.实例化()
_静态全局序号生成器 = 线程序号类.实例化(线程间独立=False)


def 序号_重置(下一个序号数字=1):
    _静态序号生成器.序号_重置(下一个序号数字)


def 序号(字符串模板="(1)"):
    return _静态序号生成器.序号(字符串模板)


def 全局序号_重置(下一个序号数字=1):
    _静态全局序号生成器.序号_重置(下一个序号数字)


def 全局序号(字符串模板="(1)"):
    return _静态全局序号生成器.序号(字符串模板)


# endregion 5.线程序号


# region 4.打点计时

# region --转换秒数相关
_毫秒_秒数 = 0.001
_秒_秒数 = _毫秒_秒数 * 1000
_分钟_秒数 = _秒_秒数 * 60
_小时_秒数 = _分钟_秒数 * 60
_天_秒数 = _小时_秒数 * 24
def _拆解秒数(秒数, 各时间单位值字典={}):
    if 秒数 > _天_秒数 + _小时_秒数:
        除余结果 = divmod(秒数, _天_秒数)
        各时间单位值字典["天"] = int(除余结果[0])
        return "%d 天, %s" % (int(除余结果[0]), _拆解秒数(除余结果[1], 各时间单位值字典))

    elif 秒数 > _小时_秒数 + _分钟_秒数:
        除余结果 = divmod(秒数, _小时_秒数)
        各时间单位值字典["小时"] = int(除余结果[0])
        return '%d 小时, %s' % (int(除余结果[0]), _拆解秒数(除余结果[1], 各时间单位值字典))

    elif 秒数 > _分钟_秒数 + _秒_秒数:
        除余结果 = divmod(秒数, _分钟_秒数)
        各时间单位值字典["分钟"] = int(除余结果[0])
        return '%d 分钟, %s' % (int(除余结果[0]), _拆解秒数(除余结果[1], 各时间单位值字典))

    elif 秒数 > _秒_秒数 + _毫秒_秒数:
        除余结果 = divmod(秒数, _秒_秒数)
        各时间单位值字典["秒"] = int(除余结果[0])
        return '%d 秒, %s' % (int(除余结果[0]), _拆解秒数(除余结果[1], 各时间单位值字典))

    else:
        除余结果 = divmod(秒数, _毫秒_秒数)
        各时间单位值字典["毫秒"] = int(除余结果[0])
        return "%d 毫秒" % int(除余结果[0])
# endregion --转换秒数相关

class 打点计时类:
    class 时间值存储类:
        def __init__(self, 时间值=0):
            self.时间值 = 时间值
            self.各时间单位值字典 = {}

        @staticmethod
        def 实例化():
            return 打点计时类.时间值存储类()

        def 可视化时间(self):
            # 清空字典
            self.各时间单位值字典["天"] = 0
            self.各时间单位值字典["小时"] = 0
            self.各时间单位值字典["分钟"] = 0
            self.各时间单位值字典["秒"] = 0
            self.各时间单位值字典["毫秒"] = 0
            return _拆解秒数(self.时间值, self.各时间单位值字典)

        def 秒数(self):
            return self.时间值

        def 最近计算的可视化时间字典(self):
            return self.各时间单位值字典

        def __str__(self):
            return self.可视化时间()

    def __init__(self, 数组型打点上限=150, 删除区间=[5, -5]):
        self.默认打点数组 = []
        self.个性打点字典 = {}
        self.时间值存储实例 = 打点计时类.时间值存储类.实例化()
        self.数组型打点上限 = 数组型打点上限
        self.删除区间 = 删除区间

    @staticmethod
    def 实例化(数组型打点上限=150, 删除区间=[5, -5]):
        return 打点计时类(数组型打点上限, 删除区间)

    def 打点(self, 计时点名称=None):
        计时点 = time.time()
        if len(self.默认打点数组) > self.数组型打点上限:
            self.默认打点数组 = self.默认打点数组[0:self.删除区间[0]] + self.默认打点数组[self.删除区间[1]:]
        self.默认打点数组.append(计时点)
        if 计时点名称:
            计时点名称 = str(计时点名称)
            self.个性打点字典[计时点名称] = 计时点

    def 计时(self, 起始点=None, 结束点=None):
        def 无参_无参_处理():
            所需计时点_数组 = self.默认打点数组[-2:]
            时间差值 = 所需计时点_数组[0] - 所需计时点_数组[-1]
            if 时间差值 < 0:
                时间差值 = -时间差值
            self.时间值存储实例.时间值 = 时间差值
            return self.时间值存储实例

        def 下标_下标_处理():
            所需计时点_数组 = self.默认打点数组[起始点:结束点]
            时间差值 = 所需计时点_数组[0] - 所需计时点_数组[-1]
            if 时间差值 < 0:
                时间差值 = -时间差值
            self.时间值存储实例.时间值 = 时间差值
            return self.时间值存储实例

        def 下标_无参_处理():
            所需计时点_数组 = self.默认打点数组[起始点:]
            时间差值 = 所需计时点_数组[0] - 所需计时点_数组[-1]
            if 时间差值 < 0:
                时间差值 = -时间差值
            self.时间值存储实例.时间值 = 时间差值
            return self.时间值存储实例

        def 名称_名称_处理():
            时间差值 = self.个性打点字典[起始点] - self.个性打点字典[结束点]
            if 时间差值 < 0:
                时间差值 = -时间差值
            self.时间值存储实例.时间值 = 时间差值
            return self.时间值存储实例

        def default():
            raise Exception("计时失败,参数类型未支持")

        switch = {
            "<class 'NoneType'><class 'NoneType'>": 无参_无参_处理,
            "<class 'int'><class 'int'>": 下标_下标_处理,
            "<class 'int'><class 'NoneType'>": 下标_无参_处理,
            "<class 'str'><class 'str'>": 名称_名称_处理
        }

        try:
            return switch.get(repr(type(起始点)) + repr(type(结束点)), default)()
        except Exception as e:
            print("计时出错")
            print(e)


_静态计时器 = 打点计时类.实例化()

def 打点(计时点名称=None):
    _静态计时器.打点(计时点名称)

def 计时(起始点=None, 结束点=None):
    return _静态计时器.计时(起始点, 结束点)

def 计时_print(起始点=None, 结束点=None, is多打个点=True):
    if is多打个点:
        打点()
    print(f"耗时: {计时(起始点,结束点)}")


# region --装饰器
def 打点计时(func):
    @wraps(func)  # 复制原始函数信息，并保留下来
    def inner(*args, **kwargs):  # args和kwargs，是原始函数的参数；args是元祖，kwargs是字典

        # region 执行原始函数前
        计时器 = 打点计时类.实例化()
        计时器.打点()
        # endregion

        rst = func(*args, **kwargs)  # 执行原始函数

        # region 执行原始函数后
        计时器.打点()
        print(f'''{func.__name__}: {计时器.计时()}''')
        # endregion

        return rst

    return inner
# endregion --装饰器


#region --可视化打点计时
_is可视化打点计时结束=True
_可视化打点计时_计时器 = 打点计时类.实例化()

def _每x行取第y行_生成器类(x, y):
    行数 = -1 - (y - 1)
    while True:
        行数 += 1
        if 行数 % x == 0:
            yield True
        else:
            yield False

def 计时点_生成器类(几个点一组=3, 几个组换行=5, 几个行成块=4, 点样式1="·", 点样式2="*"):
    每x行取第x行 = _每x行取第y行_生成器类(几个点一组, 几个点一组)
    每y行取第y行 = _每x行取第y行_生成器类(几个点一组 * 几个组换行, 几个点一组 * 几个组换行)
    每z行取第z行 = _每x行取第y行_生成器类(几个点一组 * 几个组换行 * 几个行成块, 几个点一组 * 几个组换行 * 几个行成块)
    is点样式用1 = True
    输出的点 = 点样式1
    while True:
        最终输出 = 输出的点
        if next(每x行取第x行):
            最终输出 += " "
        if next(每y行取第y行):
            最终输出 += "\n"
        if next(每z行取第z行):
            # 最终输出 += "\n"
            is点样式用1 = not is点样式用1
            if is点样式用1:
                输出的点 = 点样式1
            else:
                输出的点 = 点样式2
        yield 最终输出

def _main可视化打点计时(一个点几秒=10, 几个点一组=3, 几个组换行=4, 几个行成块=5, 点样式1="·", 点样式2="*"):
    global _is可视化打点计时结束
    _is可视化打点计时结束 = False
    _可视化打点计时_计时器.打点()
    计时点 = 计时点_生成器类(几个点一组, 几个组换行, 几个行成块, 点样式1, 点样式2)

    点数值 = 一个点几秒
    点单位 = '秒'
    行数值 = 点数值 * 几个点一组 * 几个组换行
    行单位 = '秒'
    块数值 = 行数值 * 几个行成块
    块单位 = '秒'
    if 行数值 >= 60:
        行数值 = math.floor(行数值/60) if math.floor(行数值/60)==行数值/60 else 行数值/60
        行单位 = '分钟'
    if 块数值 >= 60:
        块数值 = math.floor(块数值/60) if math.floor(块数值/60)==块数值/60 else 块数值/60
        块单位 = '分钟'
    print(f"### 一个点代表{点数值}{点单位}，一行{行数值}{行单位}，一块{块数值}{块单位} ###")

    while not _is可视化打点计时结束:
        delay_x_s(一个点几秒)
        if not _is可视化打点计时结束:
            print(next(计时点), end="")

def start可视化打点计时(一个点几秒=10, 几个点一组=3, 几个组换行=4, 几个行成块=5, 点样式1="·", 点样式2="*"):
    t1 = threading.Thread(target=_main可视化打点计时, args=(一个点几秒, 几个点一组, 几个组换行, 几个行成块, 点样式1, 点样式2))
    t1.start()

def end可视化打点计时():
    global _is可视化打点计时结束
    _is可视化打点计时结束 = True
    _可视化打点计时_计时器.打点()
    print(f"\n### 耗时: {_可视化打点计时_计时器.计时()}")
#endregion --可视化打点计时

# endregion 4.打点计时


# region 3.随机延时

# 固定延时x秒
def _delay_x_0_s(fixed_delay_num):
    x = float(fixed_delay_num)
    time.sleep(x)

# 随机延时 0~y 秒
def _delay_0_y_s(random_delay_num):
    y = float(random_delay_num)
    time.sleep(random.random() * y)


# 先固定延时x秒，再随机延时 0~y 秒
# 延时区间，包前不包后
def delay_x_y_s(fixed_delay_num, random_delay_num):
    _delay_x_0_s(fixed_delay_num)
    _delay_0_y_s(random_delay_num)


# 随机延时 x~y 秒
# 延时区间，包前不包后
def delay_x_to_y_s(最小秒数, 最大秒数):
    x = float(最小秒数)
    y = float(最大秒数)
    _delay_x_0_s(x)
    _delay_0_y_s(y - x)

delay_between_x_y_s = functools.partial(delay_x_to_y_s)


def delay_x_s(固定延时几秒):
    _delay_x_0_s(固定延时几秒)


def delay_y_s(随机延时0到几秒):
    _delay_0_y_s(随机延时0到几秒)

# endregion 3.随机延时


# region 2.数据集合

# 将lists转换成dicts
def from_rows_to_lines(源rows_lists, 标题行_list):
    rows = 源rows_lists
    col_names = 标题行_list
    lines = []
    for row in rows:
        r_dict = {}
        for i, col in enumerate(row):
            r_dict[col_names[i]] = col
        lines.append(r_dict)
    return lines



def split_list_by_count(源list, count_几个元素分一组):
    return __list_of_groups(源list, count_几个元素分一组)
def __list_of_groups(list_info, per_list_len):
    '''
    :param list_info:   列表
    :param per_list_len:  每个小列表的长度
    :return:
    '''
    list_of_group = zip(*(iter(list_info),) *per_list_len)
    end_list = [list(i) for i in list_of_group] # i is a tuple
    count = len(list_info) % per_list_len
    end_list.append(list_info[-count:]) if count !=0 else end_list
    return end_list



# 删除lsit中的某项
def delListItem(数据源list, 下标列表=None, 下标从0开始=True, 元素值列表=None, 不改变原数组=True):
    if 不改变原数组:
        数据源list = to_self(数据源list)

    if 下标列表:
        if isinstance(下标列表, str):
            下标列表 = int(下标列表)
        if isinstance(下标列表, int):
            下标列表 = [下标列表]
        if not 下标从0开始:
            下标列表 = stream(下标列表).map(lambda i: int(i) - 1).collect()
        下标列表.sort(key=None, reverse=True)
        for i in 下标列表:
            数据源list.pop(i)

    if 元素值列表:
        if not isinstance(元素值列表, (list, tuple)):
            元素值列表 = [元素值列表]
        for i in 元素值列表:
            if i in 数据源list:
                数据源list.remove(i)

    return 数据源list

# 删除多重lsit中的某项
def delMultListItem(数据源list, 多层下标字符串_list=None, 下标从0开始=True, 元素值列表=None, 不改变原数组=True, 下标分隔符="."):
    if 不改变原数组:
        数据源list = to_self(数据源list)

    if 多层下标字符串_list:
        if isinstance(多层下标字符串_list, str):
            多层下标字符串_list = [多层下标字符串_list]
        for 多层序号字符串 in 多层下标字符串_list:
            序号list = 多层序号字符串.split(下标分隔符)
            临时list = 数据源list
            for i in 序号list[:-1]:
                if 下标从0开始:
                    临时list = 临时list[int(i)]
                else:
                    临时list = 临时list[int(i) - 1]
            delListItem(临时list, 序号list[-1], 下标从0开始, 元素值列表=None, 不改变原数组=False)

    if 元素值列表:
        if not isinstance(元素值列表, (list, tuple)):
            元素值列表 = [元素值列表]
        for 元素值 in 元素值列表:
            临时list = 数据源list

            def 递归删除list中的指定元素(数据源list, 元素值):
                delListItem(数据源list, None, 下标从0开始, 元素值列表=[元素值], 不改变原数组=False)
                for i in 数据源list:
                    if isinstance(i, (list, tuple)):
                        递归删除list中的指定元素(i, 元素值)

            递归删除list中的指定元素(临时list, 元素值)

    return 数据源list



# 获取多层dict的值
def getDictValue(my_dict, key="", default=None, 分隔符="."):
    if not key:
        return default

    try:
        start_index = 0
        end_index = len(key) - 1
        if key[0] == 分隔符: start_index += 1
        if key[end_index] == 分隔符: end_index -= 1
        key = key[start_index:end_index + 1]
        keys = key.split(分隔符)
        for key in keys:
            if isinstance(my_dict, (list, tuple)):
                my_dict = my_dict[int(key)]
            else:
                my_dict = my_dict[key]
        return my_dict
    except:
        return default

# 设置多层dict的值
def setDictValue(my_dict, key, value, 分隔符='.'):
    keys = key.split(分隔符)
    length = len(keys)
    for index, i in enumerate(key.split(分隔符)):
        if int(index) + 1 == length:
            if isinstance(my_dict, (list, tuple)):
                my_dict[int(i)] = value
            else:
                my_dict[i] = value
        else:
            if isinstance(my_dict, (list, tuple)):
                my_dict = my_dict[int(i)]
            else:
                my_dict = my_dict[i]

# endregion 2.数据集合


# region 1.流式计算
from functools import cmp_to_key
from itertools import groupby


class ListStream:
    def __init__(self, my_list=[]):
        self.list = list(my_list)

    def filter(self, func):
        self.list = list(filter(func, self.list))
        return self

    def map(self, func):
        self.list = list(map(func, self.list))
        return self

    def forEach(self, func):
        list(map(func, self.list))
        return self

    def print(self):
        self.forEach(lambda item: print(item))
        return self

    def collect(self):
        return self.list

    def sort(self, key=None, reverse=False, 比较函数_返回单个值用于比较=None, 比较函数_输入两个值返回比较结果=None):
        if 比较函数_返回单个值用于比较:
            key = 比较函数_返回单个值用于比较
        elif 比较函数_输入两个值返回比较结果:
            key = cmp_to_key(比较函数_输入两个值返回比较结果)

        self.list.sort(key=key, reverse=reverse)
        return self

    def group_by_count(self, count):
        # return split_list_by_count(self.list, count)
        return [self.list[i:i+count] for i in range(0,len(self.list),count)]
    def group(self, group_key=None, sort_key=None):
        key1 = group_key
        if sort_key:
            key1 = sort_key
        self.sort(key1)

        new_list = []
        for field, group in groupby(self.list, group_key):
            new_list.append(list(group))
        return new_list
    # def group_by_cmp(self, cmp, reverse=False):
    #     # self.sort_by_cmp(cmp, reverse)
    #     new_list = []
    #     for key, group in groupby(self.list, cmp):
    #         new_list.append(list(group))
    #     return new_list


class DictStream(ListStream):
    def __init__(self, my_dict={}):
        self.list = self._dict_to_list(my_dict)

    def collect(self, is_to_dict=True):
        if is_to_dict:
            return self._list_to_dict(self.list)
        else:
            return self.list

    def _dict_to_list(self, old_dict):
        new_list = []
        for i in old_dict.keys():
            temp_dict = {}
            temp_dict["key"] = i
            temp_dict["value"] = old_dict[i]
            new_list.append(temp_dict)
        return new_list

    def _list_to_dict(self, old_list):
        new_dict = {}
        for i in old_list:
            new_dict[i["key"]] = i["value"]
        return new_dict


def stream(iteration):
    def list_处理():
        return ListStream(iteration)

    def dict_处理():
        return DictStream(iteration)

    def default():
        raise Exception("stream化失败,参数类型未支持")

    switch = {
        list: list_处理,
        tuple: list_处理,
        str: list_处理,
        dict: dict_处理
    }
    return switch.get(type(iteration), default)()

# endregion 1.流式计算
